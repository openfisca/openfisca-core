from typing import IO

import io
from datetime import date, datetime

import openpyxl
import pytest

from openfisca_core.parameters.parameter_scale import ParameterScale
from openfisca_core.reforms.reform_excel import (
    ReformExcel,
    ReformExcelBuilder,
    ReformExcelTemplateGenerator,
    get_parameter_node,
)


def to_wb_file(wb: openpyxl.Workbook) -> IO[bytes]:
    bytes = io.BytesIO()
    wb.save(bytes)
    bytes.seek(0)
    return bytes


@pytest.fixture
def wb():
    wb = openpyxl.Workbook()
    config_sheet = wb.create_sheet("Config")
    wb.create_sheet("Paramètres")
    wb.create_sheet("OtherSheet")
    config_sheet["A2"] = "root"
    config_sheet["B2"] = "benefits"

    param_sheet = wb.create_sheet("Paramètres_2025")
    param_sheet["A1"] = "Nom"
    param_sheet["B1"] = "Date"
    param_sheet["C1"] = "Valeur"

    param_sheet.append(["parenting_allowance.amount", date(2025, 1, 1), 100000])

    return wb


@pytest.fixture
def wb_file(wb: openpyxl.Workbook) -> IO[bytes]:
    return to_wb_file(wb)


class TestReformExcelBuilder:
    def test_get_suffixes(self, wb_file):
        builder = ReformExcelBuilder(baseline=None, path_or_file=wb_file)
        suffixes = builder.suffixes
        assert suffixes == ["", "_2025"]

    def test_get_reform_data(self, wb_file):
        builder = ReformExcelBuilder(baseline=None, path_or_file=wb_file)

        assert builder.root_name == "benefits"
        assert builder.get_parameters("_2025") == [
            ("parenting_allowance.amount", 100000, date(2025, 1, 1))
        ]

    def test_build_reform(self, tax_benefit_system, wb_file):
        builder = ReformExcelBuilder(baseline=tax_benefit_system, path_or_file=wb_file)
        reform = builder.build_reform("_2025")

        assert reform.root_name == "benefits"
        assert reform.reformed_parameters == [
            ("parenting_allowance.amount", 100000, date(2025, 1, 1))
        ]

    def test_pas_de_root(self, tax_benefit_system, wb):
        ws = wb["Config"]
        ws.delete_rows(2)

        builder = ReformExcelBuilder(
            baseline=tax_benefit_system, path_or_file=to_wb_file(wb)
        )
        with pytest.raises(AssertionError):
            builder.build_reform("_2025")

    def test_pas_de_period(self, tax_benefit_system, wb: openpyxl.Workbook):
        ws = wb["Paramètres_2025"]
        ws.delete_cols(1)

        builder = ReformExcelBuilder(
            baseline=tax_benefit_system, path_or_file=to_wb_file(wb)
        )
        with pytest.raises(AssertionError):
            builder.build_reform("_2025")

    def test_get_parameters_shuffled_columns(self, wb):
        param_sheet = wb["Paramètres_2025"]
        param_sheet.move_range("B1:B20", cols=2, rows=0)
        param_sheet.delete_cols(2)

        builder = ReformExcelBuilder(baseline=None, path_or_file=to_wb_file(wb))
        assert builder.get_parameters("_2025") == [
            ("parenting_allowance.amount", 100000, date(2025, 1, 1))
        ]

    def test_get_parameters_with_two_periods(self, wb: openpyxl.Workbook):
        param_sheet = wb["Paramètres_2025"]
        param_sheet["A3"] = param_sheet["A2"].value
        param_sheet["B3"] = date(2026, 1, 1)
        param_sheet["C3"] = 200000

        builder = ReformExcelBuilder(baseline=None, path_or_file=to_wb_file(wb))
        assert builder.get_parameters("_2025") == [
            ("parenting_allowance.amount", 100000, date(2025, 1, 1)),
            ("parenting_allowance.amount", 200000, date(2026, 1, 1)),
        ]


class TestReformExcel:
    def test_init(self, tax_benefit_system):
        reform = ReformExcel(
            name="name",
            baseline=tax_benefit_system,
            root_name="benefits",
            reformed_parameters=[
                ("parenting_allowance.amount", 1000.0, date(2025, 1, 1))
            ],
        )

        assert reform.root_name == "benefits"
        assert reform.reformed_parameters == [
            ("parenting_allowance.amount", 1000.0, date(2025, 1, 1))
        ]

        parameter, _ = get_parameter_node(
            reform.parameters, "benefits.parenting_allowance.amount"
        )
        assert parameter.get_at_instant("2025-01-01") == 1000.0

    def test_partials_parameters(self, tax_benefit_system):
        reform = ReformExcel(
            name="name",
            baseline=tax_benefit_system,
            root_name="benefits",
            reformed_parameters=[
                ("basic_income", 800.0, date(2026, 1, 1)),
            ],
        )

        parameter, _ = get_parameter_node(reform.parameters, "benefits.basic_income")
        assert parameter.get_at_instant("2024-06-01") == 600.0
        assert parameter.get_at_instant("2025-06-01") == 600.0
        assert parameter.get_at_instant("2026-06-01") == 800.0

        parameter2, _ = get_parameter_node(
            reform.parameters, "benefits.parenting_allowance.amount"
        )
        assert parameter2.get_at_instant("2024-06-01") == 600.0
        assert parameter2.get_at_instant("2025-06-01") == 600.0

    def test_brackets(self, tax_benefit_system):
        reform = ReformExcel(
            name="name",
            baseline=tax_benefit_system,
            root_name="taxes",
            reformed_parameters=[
                ("social_security_contribution.0", 0.01, date(2025, 1, 1)),
                ("social_security_contribution.3000", 0.03, date(2025, 1, 1)),
                ("social_security_contribution.6000", 0.13, date(2025, 1, 1)),
                ("social_security_contribution.12400", 0.001, date(2025, 1, 1)),
            ],
        )
        parameter, _ = get_parameter_node(
            reform.parameters, "taxes.social_security_contribution"
        )
        assert isinstance(parameter, ParameterScale)
        p_instant = parameter.get_at_instant("2025-01")
        assert len(p_instant.rates) == 4
        assert len(p_instant.thresholds) == 4
        assert p_instant.rates == [0.01, 0.03, 0.13, 0.001]
        assert p_instant.thresholds == [0.0, 3000.0, 6000.0, 12400.0]

    def test_unsorted_brackets(self, tax_benefit_system):
        reform = ReformExcel(
            name="name",
            baseline=tax_benefit_system,
            root_name="taxes",
            reformed_parameters=[
                ("social_security_contribution.6000", 0.13, date(2025, 1, 1)),
                ("social_security_contribution.0", 0.01, date(2025, 1, 1)),
                ("social_security_contribution.12400", 0.001, date(2025, 1, 1)),
                ("social_security_contribution.3000", 0.03, date(2025, 1, 1)),
            ],
        )
        parameter, _ = get_parameter_node(
            reform.parameters, "taxes.social_security_contribution"
        )
        assert isinstance(parameter, ParameterScale)
        p_instant = parameter.get_at_instant("2025-01")
        assert len(p_instant.rates) == 4
        assert len(p_instant.thresholds) == 4
        assert p_instant.rates == [0.01, 0.03, 0.13, 0.001]
        assert p_instant.thresholds == [0.0, 3000.0, 6000.0, 12400.0]

    def test_parameters_with_date(self, tax_benefit_system):
        reform = ReformExcel(
            name="name",
            baseline=tax_benefit_system,
            root_name="benefits",
            reformed_parameters=[
                ("basic_income", 500.0, date(2025, 1, 1)),
                ("basic_income", 600.0, date(2026, 1, 1)),
            ],
        )

        parameter, _ = get_parameter_node(reform.parameters, "benefits.basic_income")
        assert parameter.get_at_instant("2025-06-01") == 500.0
        assert parameter.get_at_instant("2026-06-01") == 600.0

    def test_bracket_parameters_with_date(self, tax_benefit_system):
        reform = ReformExcel(
            name="name",
            baseline=tax_benefit_system,
            root_name="taxes",
            reformed_parameters=[
                ("social_security_contribution.0", 0.02, date(2025, 1, 1)),
                ("social_security_contribution.6000", 0.06, date(2025, 1, 1)),
                ("social_security_contribution.12400", 0.12, date(2025, 1, 1)),
                ("social_security_contribution.0", 0.03, date(2026, 1, 1)),
                ("social_security_contribution.6000", 0.07, date(2026, 1, 1)),
                ("social_security_contribution.12400", 0.13, date(2026, 1, 1)),
            ],
        )

        parameter, _ = get_parameter_node(
            reform.parameters, "taxes.social_security_contribution"
        )
        assert isinstance(parameter, ParameterScale)

        p_instant_2025 = parameter.get_at_instant("2025-06")
        assert p_instant_2025.rates == [0.02, 0.06, 0.12]

        p_instant_2026 = parameter.get_at_instant("2026-06")
        assert p_instant_2026.rates == [0.05, 0.13, 0.25]


class TestReformExcelTemplateGenerator:
    def test_generate_parameter_data(self, tax_benefit_system):
        generator = ReformExcelTemplateGenerator(tax_benefit_system, "benefits")
        parameters = generator.parameter_data()
        assert parameters == [
            ("basic_income", 600.0),
            ("housing_allowance", None),
            ("parenting_allowance.amount", 600.0),
            ("parenting_allowance.income_threshold", 500),
        ]

    def test_generate_parameter_data_with_scale(self, tax_benefit_system):
        root_name = "taxes"
        generator = ReformExcelTemplateGenerator(tax_benefit_system, root_name)

        io_bytes = io.BytesIO()
        generator.save_template_xlsx(io_bytes)

        io_bytes.seek(0)
        wb = openpyxl.load_workbook(io_bytes)

        config_sheet = wb["Config"]
        assert config_sheet["A2"].value == "root"
        assert config_sheet["B2"].value == root_name

        param_sheet = wb["Paramètres"]
        rows = list(param_sheet.iter_rows(values_only=True))
        valeur_date = datetime(date.today().year, 1, 1)
        assert rows == [
            ("Nom", "Date", "Valeur"),
            ("housing_tax.minimal_amount", valeur_date, 200),
            ("housing_tax.rate", valeur_date, 10),
            ("income_tax_rate", valeur_date, 0.15),
            ("social_security_contribution.0.0", valeur_date, 0.02),
            ("social_security_contribution.12400.0", valeur_date, 0.12),
            ("social_security_contribution.6000.0", valeur_date, 0.06),
        ]


class TestReformExcelIntegration:
    def test_build_reform_excel_and_generate_template(self, tax_benefit_system):
        root_name = "taxes"
        generator = ReformExcelTemplateGenerator(tax_benefit_system, root_name)

        io_bytes = io.BytesIO()
        generator.save_template_xlsx(io_bytes)

        io_bytes.seek(0)

        builder = ReformExcelBuilder(baseline=tax_benefit_system, path_or_file=io_bytes)
        assert builder.suffixes == [""]

        reform = builder.build_reform(builder.suffixes[0])
        assert reform.root_name == root_name
