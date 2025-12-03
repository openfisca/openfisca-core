from __future__ import annotations

from typing import IO, Iterator

from openfisca_core.types import TaxBenefitSystem

from datetime import date
from pathlib import Path

import openpyxl

from openfisca_core.parameters.parameter_node import ParameterNode
from openfisca_core.parameters.parameter_scale import ParameterScale
from openfisca_core.parameters.parameter_scale_bracket import ParameterScaleBracket
from openfisca_core.reforms import Reform


def get_parameter_node(
    p: ParameterNode, segments: str | list[str]
) -> tuple[ParameterNode | ParameterScale | ParameterScaleBracket, list[str]]:
    if isinstance(segments, str):
        segments = segments.split(".")
    if segments:
        if hasattr(p, "children"):
            return get_parameter_node(p.children[segments[0]], segments[1:])
        else:
            return p, segments
    else:
        return p, []


class ReformExcelBuilder:
    def __init__(
        self,
        baseline: type[TaxBenefitSystem],
        path_or_file: Path | str | IO[bytes],
    ) -> None:
        # data_only = True to get the value stored the last time Excel read the sheet
        self.wb = openpyxl.load_workbook(path_or_file, data_only=True)
        self.baseline = baseline

    @property
    def suffixes(self) -> list[str]:
        return [
            sheetname[len("Paramètres") :]
            for sheetname in self.wb.sheetnames
            if sheetname.startswith("Paramètres")
        ]

    @property
    def reforms(self) -> Iterator[ReformExcel]:
        for suffix in self.suffixes:
            yield self.build_reform(suffix)

    @property
    def root_name(self) -> str:
        assert self.wb["Config"]["A2"].value == "root"
        return self.wb["Config"]["B2"].value

    def get_parameters(self, suffix: str) -> list[tuple[str, float, date]]:
        sheet = self.wb[f"Paramètres{suffix}"]
        nom_col = [
            i for i, cell in enumerate(sheet.iter_cols()) if cell[0].value == "Nom"
        ]
        valeur_col = [
            i for i, cell in enumerate(sheet.iter_cols()) if cell[0].value == "Valeur"
        ]
        date_col = [
            i for i, cell in enumerate(sheet.iter_cols()) if cell[0].value == "Date"
        ]

        assert nom_col, "La colonne 'Nom' est manquante"
        assert valeur_col is not None, "La colonne 'Valeur' est manquante"
        assert date_col, "La colonne 'Date' est manquante"

        return [
            (
                row[nom_col[0]],
                row[valeur_col[0]],
                row[date_col[0]].date(),
            )
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[date_col[0]]
        ]

    def build_reform(self, suffix: str) -> "ReformExcel":
        return ReformExcel(
            suffix,
            self.baseline,
            self.root_name,
            self.get_parameters(suffix),
        )


class ReformExcelTemplateGenerator:
    def __init__(
        self, baseline: type[TaxBenefitSystem], root_name: str, period: date
    ) -> None:
        self.baseline = baseline
        self.root_name = root_name
        self.period = period

    def generate_parameter_tree_values(self, parameter) -> list[tuple[str, float]]:
        values = []
        if type(parameter) is ParameterNode:
            for child in parameter.children.values():
                values.extend(self.generate_parameter_tree_values(child))
        else:
            value = parameter.get_at_instant(self.period)
            name = parameter.name.removeprefix(self.root_name + ".")
            if type(parameter) is ParameterScale:
                threshold_values = (
                    value.amounts
                    if parameter.metadata.get("type") == "single_amount"
                    else value.rates
                )
                for threshold, val in zip(value.thresholds, threshold_values):
                    values.append((f"{name}.{threshold}", val))
            else:
                values.append((name, value))
        return sorted(values, key=lambda v: v[0])

    def parameter_data(self) -> list[tuple[str, float]]:
        root_parameter, _ = get_parameter_node(self.baseline.parameters, self.root_name)
        return self.generate_parameter_tree_values(root_parameter)

    def save_template_xlsx(
        self,
        path_or_file: Path | str | IO[bytes],
    ) -> None:
        wb = openpyxl.Workbook()
        ws_config = wb.active
        ws_config.title = "Config"
        ws_config.append(["Parameter name", "Parameter value"])
        ws_config.append(["root", self.root_name])

        wb.create_sheet(title="Paramètres")
        ws_params = wb["Paramètres"]
        ws_params.append(["Nom", "Date", "Valeur"])

        for name, value in self.parameter_data():
            ws_params.append([name, self.period, value])

        # Heuristics to adjust to content
        ws_params.column_dimensions["A"].width = 70
        ws_params.column_dimensions["B"].width = 12
        ws_params.column_dimensions["C"].width = 12
        wb.active = ws_params

        wb.save(path_or_file)


class ReformExcel(Reform):
    """A class representing a reform defined in an Excel file.

    This class extends the Reform class to provide functionality specific to reforms defined in Excel format.
    """

    def __init__(
        self,
        name: str,
        baseline: TaxBenefitSystem,
        root_name: str,
        reformed_parameters: list[tuple[str, str, date | None]] | None = None,
    ) -> None:
        """Initialize the ReformExcel instance.

        :param baseline: Baseline TaxBenefitSystem.
        :param path: Path to the Excel file defining the reform.
        :param suffix: Suffix to identify the relevant parameters sheet.
        """
        self.name = name
        self.root_name = root_name
        self.reformed_parameters = reformed_parameters or []
        super().__init__(baseline)

    def apply(self):
        def modify_parameters(local_parameters: ParameterNode) -> ParameterNode:
            # Go down the parameter tree to the parameter root of this Excel reform
            root, _ = get_parameter_node(local_parameters, self.root_name)
            # Dict to track scale-based multiline parameters
            params_with_thresholds: dict[
                str, tuple[ParameterScale, list[tuple[float, ParameterScaleBracket]]]
            ] = dict()

            for name, value, date_ in self.reformed_parameters:
                leaf, threshold = get_parameter_node(root, name)
                if type(leaf) is ParameterScale:
                    threshold_value = float(".".join(threshold))
                    prop_name = (
                        "amount"
                        if leaf.metadata.get("type") == "single_amount"
                        else "rate"
                    )
                    stripped_name = ".".join(name.split(".")[:-1])

                    bracket = ParameterScaleBracket(
                        data={
                            "threshold": {
                                date_.isoformat(): {"value": threshold_value}
                            },
                            prop_name: {date_.isoformat(): {"value": value}},
                        }
                    )

                    params_with_thresholds.setdefault(stripped_name, (leaf, list()))[
                        1
                    ].append(
                        (
                            threshold_value,
                            bracket,
                        )
                    )
                else:
                    leaf.update(start=date_, value=value)

            for leaf, threshold in params_with_thresholds.values():
                sorted_brackets = [v[1] for v in sorted(threshold, key=lambda x: x[0])]
                leaf.brackets = sorted_brackets

            return local_parameters

        self.modify_parameters(modifier_function=modify_parameters)
